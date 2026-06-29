"""Fuente de datos de respaldo: Excel export de VenApp (Línea 58).

Cuando la BD de producción NO está accesible (sin red interna / VPN apagada), el
sistema puede generar informes a partir de un **export Excel** de la colección
`reports` (el que entrega VenApp: `report_YYYYMMDDhhmmss.xlsx`, hoja "Linea 58").

Diseño: en vez de reescribir toda la capa de datos, este módulo expone una
**colección emulada** (`ExcelReportsCollection`) que imita los métodos de PyMongo
que el código realmente usa —`count_documents`, `aggregate`, `find`, `distinct`—
sobre documentos con la MISMA forma que Mongo (campos anidados `province.name`,
`extracategory.name`, `location.coordinates`, `createdAt` como datetime, etc.).
Así `data_service`, `match_geo` y `zonas_silenciosas` funcionan SIN cambios.

Es de SOLO LECTURA por construcción (no expone métodos de escritura).
"""
from __future__ import annotations

import datetime as dt
import glob
import os
import re
from functools import lru_cache
from typing import Any

# Raíz del proyecto: app/db/excel_source.py -> <root>
_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Mapeo ruta del documento (estilo Mongo) -> columnas candidatas del export, en
# ORDEN DE PREFERENCIA. Soporta dos formatos de export:
#   * el viejo  (`report_*.xlsx`, hoja "Linea 58"): acentos y el typo "Longuitud".
#   * el nuevo  (`terremoto*.csv`): columnas sin acento (Fecha/Hora, Estatus,
#     Descripcion, Reportante, Cedula, Direccion, Longitud, Numero, Organismo).
# Se usa la PRIMERA columna candidata presente en el encabezado. Las rutas con
# punto se materializan como dicts anidados {"name": ...}.
# Nota: el export nuevo colapsa la jerarquía de categorías en una sola
# "Subcategoria"; por eso `extracategory.name` (el "tipo de emergencia" que usa
# el informe) cae a "Subcategoria" cuando no existe "Categoria extra".
_COLMAP: dict[str, tuple[str, ...]] = {
    "_id": ("ID",),
    "number": ("Codigo", "Numero"),
    "sentManually": ("Enviado Manualmente",),
    "assignedTo": ("Asignado a", "Organismo"),
    "createdAt": ("Fecha de registro", "Fecha/Hora"),
    "title": ("Titulo",),
    "category": ("Categoria",),
    "subcategory.name": ("Subcategoria",),
    "extracategory.name": ("Categoria extra", "Subcategoria"),
    "additionalcategory.name": ("Categoria adicional",),
    "status": ("Estado del reporte", "Estatus"),
    "description": ("Descripción", "Descripcion"),
    "displayName": ("Nombre", "Reportante"),
    "email": ("Email",),
    "dni": ("Cédula", "Cedula"),   # 'Cédula'/'Cedula' = de la persona afectada (PII)
    "phone_number": ("Telefono",),
    "province.name": ("Estado",),
    "municipality.name": ("Municipio",),
    "parroquia.name": ("Parroquia",),
    "address": ("Dirección", "Direccion"),
}


# --------------------------------------------------------------------------- #
# Descubrimiento del archivo Excel
# --------------------------------------------------------------------------- #
def excel_path() -> str | None:
    """Ruta del export a usar: env `VENAPP_EXCEL_PATH` o el export más reciente
    (`report_*` o `terremoto*`, .xlsx/.csv) en la raíz del proyecto. Devuelve
    None si no hay ninguno."""
    env = os.environ.get("VENAPP_EXCEL_PATH")
    if env and os.path.exists(env):
        return env
    patrones = ("report_*.xlsx", "report_*.csv", "terremoto*.xlsx", "terremoto*.csv")
    candidatos = sorted(
        (p for patron in patrones for p in glob.glob(os.path.join(_ROOT, patron))),
        key=os.path.getmtime, reverse=True,
    )
    return candidatos[0] if candidatos else None


def excel_available() -> bool:
    """True si hay un Excel de respaldo disponible (y no está deshabilitado)."""
    if os.environ.get("VENAPP_EXCEL_FALLBACK", "1") == "0":
        return False
    return excel_path() is not None


# --------------------------------------------------------------------------- #
# Carga y normalización de filas -> documentos estilo Mongo
# --------------------------------------------------------------------------- #
def _set_path(doc: dict, path: str, value: Any) -> None:
    parts = path.split(".")
    cur = doc
    for p in parts[:-1]:
        cur = cur.setdefault(p, {})
    cur[parts[-1]] = value


def _to_float(v: Any) -> float | None:
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in {"true", "1", "si", "sí", "verdadero"}


def _parse_dt(v: Any) -> dt.datetime | None:
    if isinstance(v, dt.datetime):
        return v
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # ISO 8601 (export nuevo: '2026-06-29T13:46:47.736000', con 'T' y microsegundos).
    try:
        d = dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        # El corte del evento es naive (hora local Caracas); homogeneizamos para
        # poder comparar `createdAt >= EVENT_START` sin error de tz-aware vs naive.
        return d.replace(tzinfo=None)
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_coords(v: Any) -> list[float] | None:
    """'[-68.2,10.4]' (GeoJSON [lng,lat]) -> [lng, lat]."""
    if not v:
        return None
    nums = re.findall(r"-?\d+\.?\d*", str(v))
    if len(nums) >= 2:
        try:
            return [float(nums[0]), float(nums[1])]
        except ValueError:
            return None
    return None


def _row_to_doc(row: dict) -> dict:
    def clean(val: Any) -> Any:
        if val is None:
            return None
        sval = str(val).strip()
        return None if sval == "" or sval.lower() == "nan" else val

    def pick(cols: tuple[str, ...]) -> Any:
        """Primer valor (limpio) de las columnas candidatas presentes en la fila."""
        for c in cols:
            if c in row:
                return clean(row.get(c))
        return None

    doc: dict = {}
    for path, cols in _COLMAP.items():
        raw = pick(cols)
        if path == "createdAt":
            _set_path(doc, path, _parse_dt(raw))
        elif path == "sentManually":
            _set_path(doc, path, _to_bool(raw))
        else:
            _set_path(doc, path, raw)

    lat = _to_float(pick(("Latitud",)))
    lng = _to_float(pick(("Longuitud", "Longitud")))  # 'Longuitud' (typo viejo) / 'Longitud'
    doc["latitude"] = lat
    doc["longitude"] = lng
    coords = _parse_coords(pick(("Lugar de los hechos",)))
    if coords is None and lat is not None and lng is not None:
        coords = [lng, lat]
    doc["location"] = {"coordinates": coords} if coords else {}
    return doc


def _cargar_filas_csv(path: str) -> list[dict]:
    """Lee un export CSV (mismas columnas que el Excel 'Linea 58')."""
    import csv
    docs = []
    # utf-8-sig tolera BOM; el delimitador se autodetecta (coma o punto y coma).
    with open(path, encoding="utf-8-sig", newline="") as fh:
        muestra = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(muestra, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        for row in csv.DictReader(fh, dialect=dialect):
            docs.append(_row_to_doc(row))
    return docs


@lru_cache(maxsize=1)
def _cargar_docs() -> tuple[dict, ...]:
    path = excel_path()
    if not path:
        raise RuntimeError("No hay archivo de respaldo (report_*.xlsx o report_*.csv).")
    if path.lower().endswith(".csv"):
        return tuple(_cargar_filas_csv(path))

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Linea 58"] if "Linea 58" in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    # Si hay encabezados duplicados (p. ej. dos 'Cédula'), nos quedamos con el primero.
    seen: set[str] = set()
    idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h not in seen:
            idx[h] = i
            seen.add(h)
    docs = []
    for values in it:
        row = {col: (values[i] if i < len(values) else None) for col, i in idx.items()}
        docs.append(_row_to_doc(row))
    wb.close()
    return tuple(docs)


# --------------------------------------------------------------------------- #
# Mini-motor de consulta/agregación estilo Mongo (subconjunto realmente usado)
# --------------------------------------------------------------------------- #
def _get_path(doc: dict, path: str) -> Any:
    cur: Any = doc
    for part in path.split("."):
        if isinstance(cur, dict):
            cur = cur.get(part)
        else:
            return None
    return cur


def _match_cond(value: Any, cond: Any) -> bool:
    if isinstance(cond, dict) and any(str(k).startswith("$") for k in cond):
        for op, operand in cond.items():
            if op == "$options":
                continue
            if op == "$gte":
                if value is None or value < operand:
                    return False
            elif op == "$lte":
                if value is None or value > operand:
                    return False
            elif op == "$gt":
                if value is None or value <= operand:
                    return False
            elif op == "$lt":
                if value is None or value >= operand:
                    return False
            elif op == "$ne":
                if value == operand:
                    return False
            elif op == "$in":
                if value not in operand:
                    return False
            elif op == "$nin":
                if value in operand:
                    return False
            elif op == "$regex":
                flags = re.IGNORECASE if "i" in str(cond.get("$options", "")) else 0
                if value is None or not re.search(operand, str(value), flags):
                    return False
            else:
                return False  # operador no soportado -> no coincide
        return True
    return value == cond


def _matches(doc: dict, query: dict) -> bool:
    for key, cond in query.items():
        if key == "$or":
            if not any(_matches(doc, sub) for sub in cond):
                return False
        elif key == "$and":
            if not all(_matches(doc, sub) for sub in cond):
                return False
        else:
            if not _match_cond(_get_path(doc, key), cond):
                return False
    return True


def _sort_key(field: str):
    def k(d: dict):
        v = _get_path(d, field)
        return (v is None, v)
    return k


def _eval_field(expr: Any, doc: dict) -> Any:
    if isinstance(expr, str) and expr.startswith("$"):
        return _get_path(doc, expr[1:])
    return expr


def _date_to_string(spec: dict, doc: dict) -> Any:
    val = _eval_field(spec.get("date"), doc)
    if not isinstance(val, dt.datetime):
        return None
    # El export 'Fecha de registro' ya viene en hora local (Caracas); no se aplica
    # desplazamiento de zona horaria para no duplicar el offset.
    return val.strftime(spec.get("format", "%Y-%m-%d"))


def _eval_group_id(id_spec: Any, doc: dict) -> Any:
    if isinstance(id_spec, str):
        return _eval_field(id_spec, doc)
    if isinstance(id_spec, dict):
        if "$dateToString" in id_spec:
            return _date_to_string(id_spec["$dateToString"], doc)
        # _id compuesto -> tupla hashable
        return tuple(sorted((k, _eval_field(v, doc)) for k, v in id_spec.items()))
    return id_spec


class _Cursor:
    """Cursor emulado: soporta .sort(field, dir) y .limit(n), iterable."""

    def __init__(self, docs: list[dict]):
        self._docs = docs

    def sort(self, field: str, direction: int = 1) -> "_Cursor":
        self._docs.sort(key=_sort_key(field), reverse=direction < 0)
        return self

    def limit(self, n: int) -> "_Cursor":
        if n and n > 0:
            self._docs = self._docs[:n]
        return self

    def __iter__(self):
        return iter(self._docs)


class ExcelReportsCollection:
    """Colección de SOLO LECTURA respaldada por el Excel; imita PyMongo."""

    name = "reports"

    def __init__(self, docs: tuple[dict, ...]):
        self._docs = docs

    # -- lectura básica --------------------------------------------------- #
    def count_documents(self, query: dict, **_) -> int:
        return sum(1 for d in self._docs if _matches(d, query))

    def estimated_document_count(self, **_) -> int:
        return len(self._docs)

    def find(self, query: dict | None = None, projection: dict | None = None, **_) -> _Cursor:
        q = query or {}
        return _Cursor([d for d in self._docs if _matches(d, q)])

    def distinct(self, field: str, query: dict | None = None, **_) -> list:
        docs = self._docs if not query else [d for d in self._docs if _matches(d, query)]
        out: list = []
        seen: set = set()
        for d in docs:
            v = _get_path(d, field)
            vals = v if isinstance(v, list) else [v]
            for x in vals:
                if x not in seen:
                    seen.add(x)
                    out.append(x)
        return out

    # -- agregación (subconjunto: $match,$group,$sort,$limit) ------------- #
    def aggregate(self, pipeline: list[dict], **_):
        docs: list[dict] = list(self._docs)
        for stage in pipeline:
            if "$match" in stage:
                q = stage["$match"]
                docs = [d for d in docs if _matches(d, q)]
            elif "$group" in stage:
                docs = self._group(docs, stage["$group"])
            elif "$sort" in stage:
                for field, direction in reversed(list(stage["$sort"].items())):
                    docs.sort(key=_sort_key(field), reverse=direction < 0)
            elif "$limit" in stage:
                n = stage["$limit"]
                docs = docs[:n] if n else docs
            elif "$project" in stage:
                pass  # los consumidores leen campos conocidos; no hace falta proyectar
            else:
                raise ValueError(f"Etapa de agregación no soportada: {list(stage)}")
        return iter(docs)

    @staticmethod
    def _group(docs: list[dict], spec: dict) -> list[dict]:
        id_spec = spec["_id"]
        accumulators = {k: v for k, v in spec.items() if k != "_id"}
        groups: dict = {}
        order: list = []
        for d in docs:
            keyval = _eval_group_id(id_spec, d)
            if keyval not in groups:
                g = {"_id": keyval}
                for f, acc in accumulators.items():
                    g[f] = 0 if "$sum" in acc else None
                groups[keyval] = g
                order.append(keyval)
            g = groups[keyval]
            for f, acc in accumulators.items():
                if "$sum" in acc:
                    inc = acc["$sum"]
                    g[f] += inc if isinstance(inc, (int, float)) else 1
                elif "$first" in acc and g[f] is None:
                    g[f] = _eval_field(acc["$first"], d)
                elif "$max" in acc:
                    val = _eval_field(acc["$max"], d)
                    if val is not None and (g[f] is None or val > g[f]):
                        g[f] = val
                elif "$min" in acc:
                    val = _eval_field(acc["$min"], d)
                    if val is not None and (g[f] is None or val < g[f]):
                        g[f] = val
        return [groups[k] for k in order]


@lru_cache(maxsize=1)
def get_excel_collection() -> ExcelReportsCollection:
    return ExcelReportsCollection(_cargar_docs())
